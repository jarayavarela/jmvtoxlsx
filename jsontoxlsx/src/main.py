""" Rescate de informacion de archivo en formato
    Semantic Location History ordenado en JSON -> Entrega un xlsx con la informacion
    considerada importante para el analisis de movilidad"""

import json
import os
from datetime import datetime
from pathlib import Path
from tkinter import filedialog
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def main():
    "Proceso principal"
    directorio_principal = Path(filedialog.askdirectory())
    lista_con_json = listado_de_archivos(directorio_principal)
    if not os.path.exists(f"{directorio_principal}\\salida"):
        os.mkdir(f"{directorio_principal}\\salida")
    hoja_activity_segment, workbook = crea_excel()
    n_fila = 2
    nuevo_diccionario = []
    for archivo_json in lista_con_json:
        act_seg = []
        pla_vis = []
        arch = os.path.splitext(os.path.basename(archivo_json))[0]
        with open(archivo_json, "r", encoding="utf-8") as archivo:
            diccionario_json = json.load(archivo)
            for i in range(len(diccionario_json["timelineObjects"])):
                if "activitySegment" in diccionario_json["timelineObjects"][i]:
                    act_seg.append(construccion_lista_actividades(diccionario_json, i))
                elif "placeVisit" in diccionario_json["timelineObjects"][i]:
                    pref_pla_vis = diccionario_json["timelineObjects"][i]["placeVisit"]
                    pla_vis_for = [
                        pref_pla_vis["location"]["latitudeE7"],
                        pref_pla_vis["location"]["longitudeE7"],
                    ]
                    pla_vis.append(pla_vis_for)
                else:
                    print("Verificar ERROR")
        n_fila = ingreso_datos_activity_segment(
            hoja_activity_segment, act_seg, arch, n_fila, nuevo_diccionario
        )
    workbook.save(f"{directorio_principal}\\salida\\resumen.xlsx")

def ingreso_datos_activity_segment(
    hoja,
    lista_actividades: list[list[Any]],
    archivo_origen: str,
    n_fila: int,
    nuevo_diccionario: list[dict],
) -> int:
    """ingreso de datos de actividades en libro excel"""
    for i, actividades in enumerate(lista_actividades):
        este, norte = coorde7(actividades[1], actividades[0])
        n_viaje = i + 1
        fila_inicial = n_fila
        hoja[f"C{n_fila}"] = norte
        hoja[f"D{n_fila}"] = este
        hoja[f"E{n_fila}"] = "ORIGEN"
        nuevo_diccionario.append(
            {
                "viaje": n_viaje,
                "modo": actividades[8],
                "confianza": actividades[9],
                "distancia": actividades[11],
                "tiempo": actividades[6],
            }
        )
        n_fila = n_fila + 1
        if actividades[10] != 0:
            for lat, lon in actividades[10]:
                este, norte = coorde7(lon, lat)
                hoja[f"C{n_fila}"] = norte
                hoja[f"D{n_fila}"] = este
                hoja[f"E{n_fila}"] = "RUTA"
                n_fila = n_fila + 1
        else:
            este = 0
            norte = 0
        este, norte = coorde7(actividades[3], actividades[2])
        hoja[f"C{n_fila}"] = norte
        hoja[f"D{n_fila}"] = este
        hoja[f"E{n_fila}"] = "DESTINO"
        tiempo_ini = datetime.fromisoformat(actividades[4])
        fecha_ini = tiempo_ini.date()
        hora_ini = tiempo_ini.time()
        tiempo_fin = datetime.fromisoformat(actividades[5])
        fecha_fin = tiempo_fin.date()
        hora_fin = tiempo_fin.time()
        for j in range(n_fila - fila_inicial + 1):
            hoja[f"A{fila_inicial+j}"] = n_viaje
            hoja[f"B{fila_inicial+j}"] = j + 1
            hoja[f"F{fila_inicial+j}"] = fecha_ini
            hoja[f"G{fila_inicial+j}"] = hora_ini
            hoja[f"H{fila_inicial+j}"] = fecha_fin
            hoja[f"I{fila_inicial+j}"] = hora_fin
            hoja[f"J{fila_inicial+j}"] = actividades[6]
            hoja[f"K{fila_inicial+j}"] = actividades[7]
            hoja[f"L{fila_inicial+j}"] = actividades[8]
            hoja[f"M{fila_inicial+j}"] = actividades[9]
            hoja[f"N{fila_inicial+j}"] = actividades[11]
            hoja[f"O{fila_inicial+j}"] = archivo_origen
        n_fila = n_fila + 1
    for row in hoja.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    return n_fila


def construccion_lista_actividades(
    dicc: list[dict],
    posicion: int,
) -> list[Any]:
    """Funcion para construir una lista con una actividad
    y todos sus detalles importantes"""
    pref_act_seg = dicc["timelineObjects"][posicion]["activitySegment"]
    delta = datetime.fromisoformat(
        pref_act_seg["duration"]["endTimestamp"]
    ) - datetime.fromisoformat(pref_act_seg["duration"]["startTimestamp"])
    duration_seg = round(delta.total_seconds(), 0)
    if "waypointPath" in pref_act_seg:
        if len(pref_act_seg["waypointPath"]["waypoints"]) != 0:
            way_point = [
                [
                    pref_act_seg["waypointPath"]["waypoints"][j]["latE7"],
                    pref_act_seg["waypointPath"]["waypoints"][j]["lngE7"],
                ]
                for j in range(len(pref_act_seg["waypointPath"]["waypoints"]))
            ]
        else:
            way_point = 0

        way_dist = (
            pref_act_seg["waypointPath"]["distanceMeters"]
            if "distanceMeters" in pref_act_seg["waypointPath"]
            else 0
        )
    elif "transitPath" in pref_act_seg:
        if len(pref_act_seg["transitPath"]["transitStops"]) != 0:
            way_point = [
                [
                    pref_act_seg["transitPath"]["transitStops"][j]["latitudeE7"],
                    pref_act_seg["transitPath"]["transitStops"][j]["longitudeE7"],
                ]
                for j in range(len(pref_act_seg["transitPath"]["transitStops"]))
            ]
        else:
            way_point = 0
        if "distanceMeters" in pref_act_seg["transitPath"]:
            way_dist = pref_act_seg["transitPath"]["distanceMeters"]
        else:
            way_dist = 0
    elif "simplifiedRawPath" in pref_act_seg:
        if len(pref_act_seg["simplifiedRawPath"]["points"]) != 0:
            way_point = [
                [
                    pref_act_seg["simplifiedRawPath"]["points"][j]["latE7"],
                    pref_act_seg["simplifiedRawPath"]["points"][j]["lngE7"],
                ]
                for j in range(len(pref_act_seg["simplifiedRawPath"]["points"]))
            ]
        else:
            way_point = 0
        if "distanceMeters" in pref_act_seg["simplifiedRawPath"]:
            way_dist = pref_act_seg["simplifiedRawPath"]["distanceMeters"]
        else:
            way_dist = 0
    else:
        way_point = 0
        way_dist = 0

    latini = (
        pref_act_seg["startLocation"]["latitudeE7"]
        if "latitudeE7" in pref_act_seg["startLocation"]
        else 0
    )
    lonini = (
        pref_act_seg["startLocation"]["longitudeE7"]
        if "longitudeE7" in pref_act_seg["startLocation"]
        else 0
    )
    latfin = (
        pref_act_seg["endLocation"]["latitudeE7"]
        if "latitudeE7" in pref_act_seg["endLocation"]
        else 0
    )
    lonfin = (
        pref_act_seg["endLocation"]["longitudeE7"]
        if "longitudeE7" in pref_act_seg["endLocation"]
        else 0
    )
    timeini = (
        pref_act_seg["duration"]["startTimestamp"]
        if "startTimestamp" in pref_act_seg["duration"]
        else 0
    )
    timefin = (
        pref_act_seg["duration"]["endTimestamp"]
        if "endTimestamp" in pref_act_seg["duration"]
        else 0
    )
    dist = round(pref_act_seg["distance"], 0) if "distance" in pref_act_seg else 0
    modo = pref_act_seg["activityType"] if "activityType" in pref_act_seg else 0
    confianza = pref_act_seg["confidence"] if "confidence" in pref_act_seg else 0
    return [
        latini,
        lonini,
        latfin,
        lonfin,
        timeini,
        timefin,
        duration_seg,
        dist,
        modo,
        confianza,
        way_point,
        round(way_dist, 0),
    ]


def crea_excel():
    """Excel base previo al ingreso de los datos de Actividades"""

    workbook = Workbook()
    hoja_act_seg = workbook.active
    hoja_act_seg.title = "activitySegment"
    for col in range(1, 16):  # Definir ancho para las primeras 15 columnas
        col_letter = get_column_letter(col)
        hoja_act_seg.column_dimensions[col_letter].width = 20
    hoja_act_seg["A1"] = "ID_VIAJE"
    hoja_act_seg["B1"] = "N_PUNTO"
    hoja_act_seg["C1"] = "LATITUD"
    hoja_act_seg["D1"] = "LONGITUD"
    hoja_act_seg["E1"] = "VIAJE"
    hoja_act_seg["F1"] = "FECHA_INI"
    hoja_act_seg["G1"] = "HORA_INI"
    hoja_act_seg["H1"] = "FECHA_FIN"
    hoja_act_seg["I1"] = "HORA_FIN"
    hoja_act_seg["J1"] = "DURACION (s)"
    hoja_act_seg["K1"] = "DISTANCIA_OD (m)"
    hoja_act_seg["L1"] = "MODO"
    hoja_act_seg["M1"] = "CONFIANZA"
    hoja_act_seg["N1"] = "DISTANCIA_WAY (m)"
    hoja_act_seg["O1"] = "ARCHIVO_ORIGEN"
    return hoja_act_seg, workbook


def coorde7(
    lone7: float,
    late7: float,
) -> tuple[float, float]:
    """Funcion para transformar los datos de coordenadas en formato gradosE7 a grados"""
    latitud = late7 / (10**7)
    longitud = lone7 / (10**7)

    return longitud, latitud


def listado_de_archivos(
    directorio_archivos: Path,
) -> list[Path]:
    "Funcion para listar todos los json de una carpeta"
    return [Path(str(archivo)) for archivo in directorio_archivos.glob("*.json")]
